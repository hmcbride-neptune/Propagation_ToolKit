import re
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import os
import sys
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd

# ---- Helpers for Excel engines ------------------------------------------------

def read_excel_with_engine(path: str) -> pd.DataFrame:
    """Read Excel file using correct engine based on extension.
    Supports .xlsx (openpyxl) and .xls (xlrd).
    """
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".xlsx":
            return pd.read_excel(path, engine="openpyxl")
        elif ext == ".xls":
            return pd.read_excel(path, engine="xlrd")
        else:
            # Try default as fallback
            return pd.read_excel(path)
    except Exception as e:
        raise RuntimeError(f"Failed to read Excel file: {e}")


# ---- Column matching logic ----------------------------------------------------

ALIASES = {
    "map": ["map", "map id", "map #", "map number", "mapno", "mapno.", "mapindex", "map_idx"],
    "description": ["description", "desc", "name", "service area", "area name", "map description"],
    "miu type": ["miu type", "miu", "endpoint type", "meter type", "device type", "transmitter type"],
    "read type": ["read type", "read", "reading type", "read method", "collection type"],
}


def _norm(text: str) -> str:
    return str(text).strip().lower().replace("_", " ").replace("-", " ")


def find_matching_column(df: pd.DataFrame, key: str):
    """Return original column name in df that best matches alias list for key.
    If none matched, return None.
    """
    normalized = { _norm(c): c for c in df.columns }
    for alias in ALIASES.get(key, []):
        alias_norm = _norm(alias)
        # Exact contains match: alias string inside normalized column name
        for n, orig in normalized.items():
            if alias_norm in n:
                return orig
    return None


# ---- Main App -----------------------------------------------------------------

class DataImportApp(tk.Toplevel):
    def _clear_all_internals(self):
        """Clear all internal files from the list and update UI."""
        self.internal_files.clear()
        self.refresh_internal_listbox()
        self.refresh_internal_dropdowns()
    def __init__(self, master=None):
        # If master is not provided, create a hidden root and make this
        # Toplevel a child of that root so the module can still run
        # standalone. If master is provided (from Main GUI), use it.
        self._own_root = False
        if master is None:
            self._own_root = True
            self._master_root = tk.Tk()
            self._master_root.withdraw()
            master = self._master_root
        super().__init__(master)
        self.title("DATA IMPORT TOOL")
        self.geometry("1100x650")
        try:
            self.state('zoomed')  # Start in full screen (maximized)
        except Exception:
            pass
        self.configure(bg="#d3d3d3")  # light grey background like screenshot

        # Data state
        self.internal_files = []  # full paths
        self.service_area_path = None
        self.service_area_df = None

        # Layout containers
        self._build_menubar()
        self._build_sidebar()
        self._build_main_table()

        # Default table rows (empty) when no service area uploaded
        self.load_default_rows()

        # Ensure that when running standalone closing the Toplevel exits
        # the hidden root mainloop as well.
        if self._own_root:
            def _on_close():
                try:
                    super().destroy()
                finally:
                    try:
                        self._master_root.destroy()
                    except Exception:
                        pass
            self.protocol("WM_DELETE_WINDOW", _on_close)

    def destroy(self):
        # Ensure we also destroy the hidden master if we created it.
        try:
            super().destroy()
        finally:
            if getattr(self, '_own_root', False):
                try:
                    self._master_root.destroy()
                except Exception:
                    pass

    # ---- UI Construction ----
    def _build_menubar(self):
        menubar = tk.Menu(self)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Add Internal CSV(s)...", command=self.add_internal_csvs)
        file_menu.add_command(label="Remove Selected Internal", command=self.remove_selected_internal)
        file_menu.add_separator()
        file_menu.add_command(label="Add Service Area Excel...", command=self.add_service_area_excel)
        file_menu.add_command(label="Clear Service Area", command=self.clear_service_area)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.destroy)
        menubar.add_cascade(label="File", menu=file_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About", command=lambda: messagebox.showinfo(
            "About", "Data Import Tool Built with Tkinter and pandas© 2025"))
        menubar.add_cascade(label="Help", menu=help_menu)

        self.config(menu=menubar)

    def on_senet_toggle(self):
            if self.senet_var.get():
                messagebox.showwarning(
                    "Senet Feature",
                    "This feature is under development. Please send your Senet files to Hagan for testing. Thanks!"
                )

    def _build_sidebar(self):
        sidebar = tk.Frame(self, bg="#e6e6e6", width=240)
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)

        # Internal Files section
        lbl_internal = tk.Label(sidebar, text="Internal Files", bg="#e6e6e6", fg="#333", font=("Segoe UI", 11, "bold"))
        lbl_internal.pack(anchor="w", padx=12, pady=(12, 6))

        # Text widget with wrapping for long file names
        int_listbox_frame = tk.Frame(sidebar, bg="#e6e6e6")
        int_listbox_frame.pack(fill=tk.X, padx=12)
        self.internal_listbox = tk.Text(int_listbox_frame, height=10, wrap=tk.CHAR,
                                        state=tk.DISABLED, cursor="arrow")
        vsb = tk.Scrollbar(int_listbox_frame, orient=tk.VERTICAL, command=self.internal_listbox.yview)
        self.internal_listbox.configure(yscrollcommand=vsb.set)
        self.internal_listbox.tag_configure("selected", background="#0078d7", foreground="white")
        self.internal_listbox.pack(fill=tk.X, side=tk.LEFT, expand=True)
        vsb.pack(fill=tk.Y, side=tk.RIGHT)
        self.internal_listbox.bind("<Button-1>", self._on_internal_click)
        self._internal_selected = set()
        int_btns = tk.Frame(sidebar, bg="#e6e6e6")
        int_btns.pack(fill=tk.X, padx=12, pady=6)
        tk.Button(int_btns, text="Add...", command=self.add_internal_csvs).pack(side=tk.LEFT)
        tk.Button(int_btns, text="Delete", command=self.remove_selected_internal).pack(side=tk.LEFT, padx=6)
        # add button to clear all
        tk.Button(int_btns, text="Clear All", command=self._clear_all_internals).pack(side=tk.LEFT, padx=6)

        # Service Area section
        lbl_sa = tk.Label(sidebar, text="Service Area File", bg="#e6e6e6", fg="#333", font=("Segoe UI", 11, "bold"))
        lbl_sa.pack(anchor="w", padx=12, pady=(0, 6))

        self.sa_file_label = tk.Label(sidebar, text="(none)", bg="#e6e6e6", fg="#555", wraplength=210, justify=tk.LEFT)
        self.sa_file_label.pack(fill=tk.X, padx=12)

        sa_btns = tk.Frame(sidebar, bg="#e6e6e6")
        sa_btns.pack(fill=tk.X, padx=12, pady=6)
        tk.Button(sa_btns, text="Add...", command=self.add_service_area_excel).pack(side=tk.LEFT)
        tk.Button(sa_btns, text="Clear", command=self.clear_service_area).pack(side=tk.LEFT, padx=6)

        # Coverage Analysis Case input
        lbl_coverage = tk.Label(sidebar, text="Coverage Analysis Case", bg="#e6e6e6", fg="#333", font=("Segoe UI", 11, "bold"))
        lbl_coverage.pack(anchor="w", padx=12, pady=(18, 6))
        self.coverage_entry = tk.Entry(sidebar, bg="#fff", fg="#222", font=("Segoe UI", 10), relief=tk.SUNKEN)
        self.coverage_entry.pack(fill=tk.X, padx=12, pady=(0, 12))

        # CCA and Submit buttons at the bottom of the sidebar
        sidebar.update_idletasks()
        btns_frame = tk.Frame(sidebar, bg="#e6e6e6")
        btns_frame.pack(fill=tk.X, padx=12, pady=(0, 16))
        self.senet_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            btns_frame,
            text="Senet",
            bg="#e6e6e6",
            variable=self.senet_var,
            # command=self.on_senet_toggle
        ).pack(side=tk.LEFT, expand=True, fill=tk.X)
        tk.Button(btns_frame, text="CCA", command=self.on_ccal1, font=("Segoe UI", 10, "bold"), bg="#2e7da8", fg="white", activebackground="#246688", activeforeground="white", bd=0, relief=tk.FLAT, padx=8, pady=4).pack(side=tk.LEFT, expand=True, fill=tk.X)
        tk.Button(btns_frame, text="Submit", command=self.on_submit, font=("Segoe UI", 10, "bold"), bg="#2e7da8", fg="white", activebackground="#246688", activeforeground="white", bd=0, relief=tk.FLAT, padx=8, pady=4).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(8,0))

        # add a button to reset all inputs to default
        def reset_all():
            self._clear_all_internals()
            self.clear_service_area()
            self.coverage_entry.delete(0, tk.END)
            self.load_default_rows()
        tk.Button(sidebar, text="Reset All", command=reset_all, font=("Segoe UI", 10, "bold"), bg="#a8a8a8", fg="white", activebackground="#888888", activeforeground="white", bd=0, relief=tk.FLAT, padx=8, pady=4).pack(side=tk.BOTTOM, fill=tk.X, padx=12, pady=(0,16))
        

    def _build_main_table(self):
        # ...existing code...
        self.main_area = tk.Frame(self, bg="#d3d3d3")
        self.main_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Table and logo split left/right
        table_frame = tk.Frame(self.main_area, bg="#d3d3d3")
        table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Title centered-ish
        title = tk.Label(table_frame, text="DATA IMPORT TOOL", bg="#d3d3d3", fg="#222", font=("Segoe UI", 14, "bold"))
        title.pack(pady=(16, 4))

        # Table frame
        self.table_frame = tk.Frame(table_frame, bg="#d3d3d3")
        self.table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(10, 30))

        self.columns = ["Map", "Description", "MIU Type", "Read Type", "Internal File"]
        style = ttk.Style()
        style.configure("Custom.Treeview", borderwidth=1, relief="solid", rowheight=24)
        style.configure("Custom.Treeview.Heading", borderwidth=1, relief="solid")
        # Add row border color for main table
        style.map("Custom.Treeview", background=[('selected', '#cce6ff')])
        style.layout("Custom.Treeview", [
            ('Treeview.field', {'sticky': 'nswe', 'children': [
                ('Treeview.padding', {'sticky': 'nswe', 'children': [
                    ('Treeview.treearea', {'sticky': 'nswe'})
                ]})
            ]})
        ])
        self.tree = ttk.Treeview(self.table_frame, columns=self.columns, show="headings", style="Custom.Treeview", selectmode='extended')
        for col in self.columns:
            self.tree.heading(col, text=col)
        self.tree.column("Map", width=80, anchor=tk.CENTER)
        self.tree.column("Description", width=240)
        self.tree.column("MIU Type", width=140)
        self.tree.column("Read Type", width=140)
        import tkinter.font as tkfont
        header_font = tkfont.nametofont("TkHeadingFont") if "TkHeadingFont" in tkfont.names() else tkfont.nametofont("TkDefaultFont")
        header_width = header_font.measure("Internal File") + 20  # add some padding
        min_width = 60
        col_width = max(header_width, min_width)
        self.tree.column("Internal File", width=col_width, anchor=tk.CENTER)
        # Add striped row tags for alternating colors (after self.tree is created)
        self.tree.tag_configure('oddrow', background='#f2f2f2')
        self.tree.tag_configure('evenrow', background='#ffffff')

        vsb = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=lambda *args: self._auto_hide_scrollbar(vsb, *args), xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.table_frame.grid_columnconfigure(0, weight=1)
        self.table_frame.grid_rowconfigure(0, weight=1)

        # Logo frame on the right, vertically centered, always present
        self.logo_frame = tk.Frame(self.main_area, bg="#d3d3d3", width=200)
        self.logo_frame.pack(side=tk.RIGHT, fill=tk.Y)
        self.logo_frame.pack_propagate(False)
        # Resolve an absolute path to the Logos folder (avoid relative-path issues)
        logo_path = os.path.join(os.path.dirname(__file__), "Logos", "NeptuneLogo.png")
        try:
            if not hasattr(self, "_logo_img"):
                # Use Pillow's ImageTk which reliably supports BMP/PNG on all platforms
                from PIL import Image, ImageTk
                img = Image.open(logo_path)
                self._logo_img = ImageTk.PhotoImage(img)
            logo_label = tk.Label(self.logo_frame, image=self._logo_img, bg="#d3d3d3")
            logo_label.place(relx=0.5, rely=0.5, anchor="center")
            self._logo_label = logo_label
        except Exception as e:
            # Fail silently in the UI but log to console for debugging
            print(f"Failed to load logo '{logo_path}': {e}")

        # --- Right-click context menu for assigning internal file ---
        self._tree_menu = tk.Menu(self, tearoff=0)
        self._tree_menu.add_command(label="Assign Internal File...", command=self._show_assign_internal_dialog)
        self.tree.bind("<Button-3>", self._on_tree_right_click)

    def _on_tree_right_click(self, event):
        # Select row under cursor if not already selected
        row_id = self.tree.identify_row(event.y)
        if row_id:
            if row_id not in self.tree.selection():
                self.tree.selection_set(row_id)
        # Show menu at cursor
        try:
            self._tree_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._tree_menu.grab_release()

    def _show_assign_internal_dialog(self):
        # Dialog to pick internal file letter and assign to all selected rows
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("No selection", "Please select one or more rows to assign.")
            return
        # Build dialog
        dialog = tk.Toplevel(self)
        dialog.title("Assign Internal File")
        dialog.geometry("300x120")
        tk.Label(dialog, text="Assign Internal File to selected rows:").pack(pady=(12, 6))
        values = [chr(ord('A') + idx) for idx in range(len(self.internal_files))]
        var = tk.StringVar()
        cb = ttk.Combobox(dialog, values=values, textvariable=var, state='readonly')
        cb.pack(pady=6)
        cb.focus_set()
        def assign_and_close():
            chosen = var.get()
            if not chosen:
                messagebox.showwarning("No selection", "Please select an internal file.")
                return
            for rid in selected:
                vals = list(self.tree.item(rid, "values"))
                if len(vals) >= 5:
                    vals[4] = chosen
                    self.tree.item(rid, values=vals)
            dialog.destroy()
        btn = tk.Button(dialog, text="Assign", command=assign_and_close)
        btn.pack(pady=(6, 12))
        dialog.transient(self)
        dialog.grab_set()
        dialog.wait_window()

    def _auto_hide_scrollbar(self, scrollbar, first, last):
        # Show scrollbar only if not all rows are visible
        try:
            first, last = float(first), float(last)
        except Exception:
            scrollbar.grid()
            return
        if last - first >= 0.999:
            scrollbar.grid_remove()
        else:
            scrollbar.grid()
        # Enable Internal File dropdown on single click
        self.tree.bind("<Button-1>", self._on_internal_file_click)
        
    def _on_internal_file_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)  # '#1'..'#5'
        col_index = int(col_id.replace('#', '')) - 1
        if col_index != 4:
            return
        bbox = self.tree.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox
        values = [chr(ord('A') + idx) for idx in range(len(self.internal_files))]
        import tkinter.font as tkfont
        font = tkfont.nametofont("TkDefaultFont")
        if values:
            max_pixel_width = max(font.measure(v) for v in values)
            pixel_width = max(w, max_pixel_width + 40)
        else:
            pixel_width = w
        cb = ttk.Combobox(self.tree, values=values, state='readonly')
        cb.place(x=x, y=y, width=pixel_width, height=h)
        current = self.tree.set(row_id, self.columns[col_index])
        if current:
            cb.set(current)
        cb.focus_set()
        cb.event_generate('<Button-1>')  # Open dropdown immediately

        # Get all selected rows, or just the clicked row if none selected
        selected = self.tree.selection()
        if not selected or row_id not in selected:
            selected = [row_id]

        def finish(_evt=None):
            chosen = cb.get()
            for rid in selected:
                self.tree.set(rid, self.columns[col_index], chosen)
            cb.destroy()

        cb.bind("<<ComboboxSelected>>", finish)
        cb.bind("<FocusOut>", lambda e: cb.destroy())

    # ---- Default / Reset ----
    def load_default_rows(self):
        self.tree.delete(*self.tree.get_children())
        for i in range(1, 7):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            self.tree.insert("", tk.END, values=[i, "", "", "", ""], tags=(tag,))  # 6 blank rows

    # ---- Internal files management ----
    def add_internal_csvs(self):
        paths = filedialog.askopenfilenames(title="Select Internal CSV files", filetypes=[("CSV files", "*.csv")])
        if not paths:
            return
        # Append unique paths
        added = 0
        for p in paths:
            if p not in self.internal_files:
                self.internal_files.append(p)
                added += 1
        if added == 0:
            return
        self.refresh_internal_listbox()
        self.refresh_internal_dropdowns()

    def _on_internal_click(self, event):
        widget = self.internal_listbox
        line = int(widget.index(f"@{event.x},{event.y}").split('.')[0])
        idx = line - 1
        if 0 <= idx < len(self.internal_files):
            if idx in self._internal_selected:
                self._internal_selected.discard(idx)
                widget.tag_remove("selected", f"{line}.0", f"{line}.end+1c")
            else:
                self._internal_selected.add(idx)
                widget.tag_add("selected", f"{line}.0", f"{line}.end+1c")

    def remove_selected_internal(self):
        selection = sorted(self._internal_selected, reverse=True)
        if not selection:
            return
        for idx in selection:
            try:
                self.internal_files.pop(idx)
            except IndexError:
                pass
        self._internal_selected = set()
        self.refresh_internal_listbox()
        self.refresh_internal_dropdowns()

    def refresh_internal_listbox(self):
        self._internal_selected = set()
        self.internal_listbox.configure(state=tk.NORMAL)
        self.internal_listbox.delete("1.0", tk.END)
        for idx, p in enumerate(self.internal_files):
            letter = chr(ord('A') + idx)
            display = f"{letter}: {os.path.basename(p)}"
            self.internal_listbox.insert(tk.END, display + "\n")
        self.internal_listbox.configure(state=tk.DISABLED)

    def refresh_internal_dropdowns(self):
        # If table already has values in the Internal File column, ensure they still exist;
        # remove any values no longer present.
        available = [chr(ord('A') + idx) for idx in range(len(self.internal_files))]
        for item in self.tree.get_children():
            vals = list(self.tree.item(item, "values"))
            if len(vals) >= 5:
                chosen = vals[4]
                if chosen and chosen not in available:
                    # Clear value if the chosen index was deleted
                    vals[4] = ""
                    self.tree.item(item, values=vals)

    # ---- Service Area management ----
    def add_service_area_excel(self):
        path = filedialog.askopenfilename(title="Select Service Area Excel", filetypes=[("Excel files", "*.xlsx;*.xls")])
        if not path:
            return
        try:
            # Read with no header, handle merged headers manually
            df_raw = pd.read_excel(path, header=None)
            # Assume first two columns are 'Map' and 'Description' (merged), next columns use row 1 (second row, 0-based)
            columns = ["Map", "Description"] + list(df_raw.iloc[1, 2:])
            df = df_raw.iloc[2:].copy()
            df.columns = columns
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return
        self.service_area_path = path
        self.service_area_df = df
        self.sa_file_label.configure(text=os.path.basename(path))
        self.populate_table_from_service_area()

    def clear_service_area(self):
        self.service_area_path = None
        self.service_area_df = None
        self.sa_file_label.configure(text="(none)")
        self.load_default_rows()

    def populate_table_from_service_area(self):
        df = self.service_area_df
        if df is None or df.empty:
            self.load_default_rows()
            return

        # Find matching columns in the sheet
        map_col = find_matching_column(df, "map")
        desc_col = find_matching_column(df, "description")
        miu_col = find_matching_column(df, "miu type")
        read_col = find_matching_column(df, "read type")

        # Filter rows to only those with a non-empty, non-NaN Description
        filtered_df = df[df[desc_col].notna() & (df[desc_col].astype(str).str.strip() != "")]

        # Sort by Map column (try numeric, fallback to string)
        if map_col:
            try:
                filtered_df = filtered_df.copy()
                filtered_df["__map_sort"] = pd.to_numeric(filtered_df[map_col], errors="coerce")
                filtered_df = filtered_df.sort_values(by="__map_sort", na_position="last")
            except Exception:
                filtered_df = filtered_df.sort_values(by=map_col, na_position="last")
        # Only clear and repopulate the tree, never destroy/recreate it
        self.tree.delete(*self.tree.get_children())

        for i in range(len(filtered_df)):
            row = filtered_df.iloc[i]
            map_val = row[map_col] if map_col else i + 1
            desc_val = row[desc_col] if desc_col else ""
            miu_val = row[miu_col] if miu_col else ""
            read_val = row[read_col] if read_col else ""
            # Convert to plain strings for display
            vals = [map_val, desc_val, miu_val, read_val, ""]
            vals = ["" if pd.isna(v) else str(v) for v in vals]
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            self.tree.insert("", tk.END, values=vals, tags=(tag,))

    # ---- Cell editing (combobox overlay for last column) ----
    def _on_cell_double_click(self, event):
        # Only allow editing for the Internal File column
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)  # '#1'..'#5'
        col_index = int(col_id.replace('#', '')) - 1
        bbox = self.tree.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox

        if col_index == 4:  # Internal File dropdown only
            values = [chr(ord('A') + idx) for idx in range(len(self.internal_files))]
            import tkinter.font as tkfont
            font = tkfont.nametofont("TkDefaultFont")
            if values:
                max_pixel_width = max(font.measure(v) for v in values)
                pixel_width = max(w, max_pixel_width + 40)
            else:
                pixel_width = w
            cb = ttk.Combobox(self.tree, values=values, state='readonly')
            cb.place(x=x, y=y, width=pixel_width, height=h)
            current = self.tree.set(row_id, self.columns[col_index])
            if current:
                cb.set(current)
            cb.focus_set()

            def finish(_evt=None):
                chosen = cb.get()
                self.tree.set(row_id, self.columns[col_index], chosen)
                cb.destroy()

            cb.bind("<<ComboboxSelected>>", finish)
            cb.bind("<FocusOut>", lambda e: cb.destroy())
        # All other columns are not editable

    # ---- Actions ----
    def on_ccal1(self):
        # --- Level 2 CCA detection ---
        is_level2_cca = False
        level2_col_name = None
        if len(self.internal_files) == 1:
            internal_path = self.internal_files[0]
            try:
                df_internal = pd.read_csv(internal_path)
                for col in df_internal.columns:
                    if col.strip().lower() == "level 2 result":
                        is_level2_cca = True
                        level2_col_name = col
                        break
            except Exception:
                pass
        # Now is_level2_cca is True if Level 2 Results column exists
        result_col = "Result"
        description_val = "CCA"
        if is_level2_cca and level2_col_name:
            result_col = level2_col_name
            description_val = "Level 2 Cellular"
        # Remove any previous Accept/Back buttons from the logo frame (if it exists)
        if hasattr(self, 'logo_frame') and self.logo_frame.winfo_exists():
            for widget in self.logo_frame.winfo_children():
                if getattr(widget, '_cca_action_button', False):
                    widget.destroy()

        # Remove the original logo frame before adding a new one
        if hasattr(self, 'logo_frame') and self.logo_frame.winfo_exists():
            self.logo_frame.destroy()
        # Recreate the logo frame on the right
        self.logo_frame = tk.Frame(self.main_area, bg="#d3d3d3", width=200)
        self.logo_frame.pack(side=tk.RIGHT, fill=tk.Y)
        self.logo_frame.pack_propagate(False)
        logo_path = os.path.join(os.path.dirname(__file__), "NeptuneLogo.png")
        try:
            if not hasattr(self, "_logo_img"):
                self._logo_img = tk.PhotoImage(file=logo_path)
            logo_label = tk.Label(self.logo_frame, image=self._logo_img, bg="#d3d3d3")
            logo_label.place(relx=0.5, rely=0.5, anchor="center")
            self._logo_label = logo_label
        except Exception as e:
            pass

        # Add Accept and Back buttons to the right side (logo_frame) after logo is created
        def on_accept():
            # Export the CCA table to CSV in Z:/Propagations/DataImport or fallback to working directory
            cca_dir = r"Z:\\Propagations\\DataImport"
            # Find the CCA table's Treeview widget
            cca_tree = None
            for widget in self._cca_table_frame.winfo_children():
                for subwidget in widget.winfo_children():
                    if isinstance(subwidget, ttk.Treeview):
                        cca_tree = subwidget
                        break
                if cca_tree:
                    break
            if cca_tree is None:
                messagebox.showerror("Error", "Could not find CCA table to export.")
                return
            # Get data from the CCA table
            columns = cca_tree.cget("columns")
            data = [columns]
            for item in cca_tree.get_children():
                row = cca_tree.item(item, "values")
                data.append(row)
            # Build output filename
            internal_path = self.internal_files[0] if self.internal_files else "internal.csv"
            base = os.path.splitext(os.path.basename(internal_path))[0]
            out_name = f"{base} Data Import.csv"
            out_path = os.path.join(cca_dir, out_name)
            try:
                os.makedirs(cca_dir, exist_ok=True)
                with open(out_path, "w", newline='', encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerows(data)
                messagebox.showinfo("Accepted", f"CCA Data exported to {out_path}")
            except Exception:
                # Fallback to working directory
                out_path = os.path.join(os.getcwd(), out_name)
                try:
                    with open(out_path, "w", newline='', encoding="utf-8") as f:
                        writer = csv.writer(f)
                        writer.writerows(data)
                    messagebox.showinfo("Accepted", f"CCA Data exported to {out_path}")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to export CCA Data: {e}")

        def on_back():
            # Hide CCA table and show main table again
            if hasattr(self, '_cca_table_frame') and self._cca_table_frame.winfo_exists():
                self._cca_table_frame.pack_forget()
            if hasattr(self, 'table_frame') and self.table_frame.winfo_exists():
                parent = self.table_frame.master
                parent.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            # Remove Accept/Back buttons frame
            if hasattr(self, '_cca_btns_frame') and self._cca_btns_frame.winfo_exists():
                self._cca_btns_frame.destroy()
            # Extra: Remove any button frames or widgets in logo_frame that are not the logo label
            if hasattr(self, 'logo_frame') and self.logo_frame.winfo_exists():
                for widget in self.logo_frame.winfo_children():
                    # Only destroy if it's a Frame or Button and not the logo label
                    if widget is not getattr(self, '_logo_label', None):
                        widget.destroy()

        # Create a frame to hold the buttons, placed below the logo and centered
        btns_frame = tk.Frame(self.logo_frame, bg="#d3d3d3")
        btns_frame.place(relx=0.5, rely=0.65, anchor="n")  # 0.65 = below center
        self._cca_btns_frame = btns_frame  # Save reference for removal

        accept_btn = tk.Button(btns_frame, text="Accept", command=on_accept, font=("Segoe UI", 10, "bold"), bg="#2e7da8", fg="white", activebackground="#246688", activeforeground="white", bd=0, relief=tk.FLAT, padx=16, pady=8)
        accept_btn._cca_action_button = True
        accept_btn.pack(side=tk.TOP, pady=(0, 12), fill=tk.X)
        back_btn = tk.Button(btns_frame, text="Back", command=on_back, font=("Segoe UI", 10, "bold"), bg="#a8a8a8", fg="white", activebackground="#888888", activeforeground="white", bd=0, relief=tk.FLAT, padx=16, pady=8)
        back_btn._cca_action_button = True
        back_btn.pack(side=tk.TOP, fill=tk.X)

        # Take the internal file. For CCA there should only be one.
        if len(self.internal_files) != 1:
            messagebox.showerror("Error", "CCA requires exactly one internal file.")
            return
        internal_path = self.internal_files[0]
        try:
            df_internal = pd.read_csv(internal_path)
            if "Result" in df_internal.columns:
                df_internal["Result"] = df_internal["Result"].replace("R900_ Wall External LoRaWAN", "R900_Wall External LoRaWAN")
            # Normalize Result column for Wall External LoRaWAN typo
            if "Result" in df_internal.columns:
                df_internal["Result"] = df_internal["Result"].replace("R900_ Wall External LoRaWAN", "R900_Wall External LoRaWAN")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read internal CSV: {e}")
            return


        # Hide the main table frame (parent of self.table_frame) but keep the logo frame visible on the right
        if hasattr(self, 'table_frame') and self.table_frame.winfo_ismapped():
            parent = self.table_frame.master
            parent.pack_forget()

        # Remove the original logo frame before adding a new one
        if hasattr(self, 'logo_frame') and self.logo_frame.winfo_exists():
            self.logo_frame.destroy()
        # Recreate the logo frame on the right
        self.logo_frame = tk.Frame(self.main_area, bg="#d3d3d3", width=200)
        self.logo_frame.pack(side=tk.RIGHT, fill=tk.Y)
        self.logo_frame.pack_propagate(False)
        logo_path = os.path.join(os.path.dirname(__file__), "NeptuneLogo.png")
        try:
            if not hasattr(self, "_logo_img"):
                self._logo_img = tk.PhotoImage(file=logo_path)
            logo_label = tk.Label(self.logo_frame, image=self._logo_img, bg="#d3d3d3")
            logo_label.place(relx=0.5, rely=0.5, anchor="center")
            self._logo_label = logo_label
        except Exception as e:
            pass

        # Create a frame to hold the buttons, placed below the logo and centered
        btns_frame = tk.Frame(self.logo_frame, bg="#d3d3d3")
        btns_frame.place(relx=0.5, rely=0.65, anchor="n")  # 0.65 = below center

        accept_btn = tk.Button(btns_frame, text="Accept", command=on_accept, font=("Segoe UI", 10, "bold"), bg="#2e7da8", fg="white", activebackground="#246688", activeforeground="white", bd=0, relief=tk.FLAT, padx=16, pady=8)
        accept_btn._cca_action_button = True
        accept_btn.pack(side=tk.TOP, pady=(0, 12), fill=tk.X)
        back_btn = tk.Button(btns_frame, text="Back", command=on_back, font=("Segoe UI", 10, "bold"), bg="#a8a8a8", fg="white", activebackground="#888888", activeforeground="white", bd=0, relief=tk.FLAT, padx=16, pady=8)
        back_btn._cca_action_button = True
        back_btn.pack(side=tk.TOP, fill=tk.X)

        # If a CCA table already exists, clear it; otherwise, create it
        if not hasattr(self, '_cca_table_frame') or not self._cca_table_frame.winfo_exists():
            self._cca_table_frame = tk.Frame(self.main_area, bg="#d3d3d3")
            self._cca_table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=(10, 30))
        else:
            for widget in self._cca_table_frame.winfo_children():
                widget.destroy()
            self._cca_table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=(10, 30))

        # New title for CCA table
        title = tk.Label(self._cca_table_frame, text="CCA DATA Import", bg="#d3d3d3", fg="#222", font=("Segoe UI", 14, "bold"))
        title.pack(pady=(16, 4))

        columns = ["Date", "Description", "Read Type", "Network/System Type", "Endpoint Type", "Collectors", "Total Services", "Invalid Coordinates", "Adequate Coverage", "Inconclusive Coverage", "Area Total", "Area Coverage", "Coverage Analysis Case"]
        style = ttk.Style()
        style.configure("CCAL.Treeview", borderwidth=1, relief="solid", rowheight=24)
        style.configure("CCAL.Treeview.Heading", borderwidth=1, relief="solid")
        style.map("CCAL.Treeview", background=[('selected', '#cce6ff')])
        style.layout("CCAL.Treeview", [
            ('Treeview.field', {'sticky': 'nswe', 'children': [
                ('Treeview.padding', {'sticky': 'nswe', 'children': [
                    ('Treeview.treearea', {'sticky': 'nswe'})
                ]})
            ]})
        ])
        # Use a frame for the table and scrollbars
        cca_table_inner = tk.Frame(self._cca_table_frame, bg="#d3d3d3")
        cca_table_inner.pack(fill=tk.BOTH, expand=True)
        tree = ttk.Treeview(cca_table_inner, columns=columns, show="headings", style="CCAL.Treeview")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor=tk.CENTER)
        vsb = ttk.Scrollbar(cca_table_inner, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(cca_table_inner, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        cca_table_inner.grid_columnconfigure(0, weight=1)
        cca_table_inner.grid_rowconfigure(0, weight=1)

        # --- Make CCA table cells editable on click ---
        def on_cca_cell_click(event):
            region = tree.identify("region", event.x, event.y)
            if region != "cell":
                return
            row_id = tree.identify_row(event.y)
            col_id = tree.identify_column(event.x)  # '#1'..'#N'
            col_index = int(col_id.replace('#', '')) - 1
            bbox = tree.bbox(row_id, col_id)
            if not bbox:
                return
            x, y, w, h = bbox
            current = tree.set(row_id, columns[col_index])
            entry = tk.Entry(tree)
            entry.place(x=x, y=y, width=w, height=h)
            entry.insert(0, current)
            entry.focus_set()

            def finish(_evt=None):
                tree.set(row_id, columns[col_index], entry.get())
                entry.destroy()

            entry.bind("<Return>", finish)
            entry.bind("<FocusOut>", lambda e: entry.destroy())

        tree.bind("<Button-1>", on_cca_cell_click)

        filename = os.path.basename(internal_path)
        date_match = re.search(r'(\d{6})', filename)
        if date_match:
            date_str = date_match.group(1)
            month = date_str[2:4]
            day = date_str[4:6]
            year = "20" + date_str[0:2]
            formatted_date = f"{month}/{day}/{year}"
        else:
            formatted_date = "01/01/2000"  # default date if not found

        endpoint_type = "R900_Cellular"
        collectors = 0
        total_services = len(df_internal)
        area_total = ""
        area_coverage = ""
        coverage_case = self.coverage_entry.get().strip()
        read_type = ""

        providers = []
        if "LTE-M_-100" in df_internal.columns:
            providers.append("FirstNet")
        if "CAT-M" in df_internal.columns:
            providers.append("Verizon")
        if "L700-CATM" in df_internal.columns or "L1900-CATM" in df_internal.columns:
            providers.append("T-Mobile")
        for provider in providers:
            if provider == "FirstNet":
                adeq_cov = df_internal[(df_internal["LTE-M_-100"].astype(str).str.upper() == "Y") | (df_internal["LTE-M_-100"].astype(str).str.upper() == "YES")].shape[0]
            elif provider == "Verizon":
                adeq_cov = df_internal[(pd.to_numeric(df_internal["CAT-M"], errors='coerce') >= -100) & (pd.to_numeric(df_internal["CAT-M"], errors='coerce') != 0)].shape[0]
            elif provider == "T-Mobile":
                adeq_cov = df_internal[(df_internal["L700-CATM"].astype(str).str.upper().isin(["IBC", "IBR"])) | (df_internal["L1900-CATM"].astype(str).str.upper().isin(["IBC", "IBR"]))].shape[0]
            else:
                adeq_cov = 0
            invalid_coords = df_internal[result_col].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"]).sum()
            inconcl_cov = total_services - adeq_cov - invalid_coords
            row_vals = [
                formatted_date,
                description_val if is_level2_cca else provider,
                read_type,
                provider,
                endpoint_type,
                collectors,
                total_services,
                invalid_coords,
                adeq_cov,
                inconcl_cov,
                area_total,
                area_coverage,
                coverage_case
            ]
            tree.insert("", tk.END, values=row_vals)

        # Add extra row with Description and Network/System Type as 'CCA' or 'Level 2 Cellular'
        # Defensive: Only calculate if result_col exists in df_internal
        if result_col in df_internal.columns:
            adequate_coverage = (
                (df_internal[result_col].astype(str).str.lower() == "firstnet") |
                (df_internal[result_col].astype(str).str.lower() == "verizon") |
                (df_internal[result_col].astype(str).str.lower() == "t-mobile")|
                (df_internal[result_col].astype(str).str.lower() == "r900_cellular")
            ).sum()
            total_inconclusive = (df_internal[result_col].astype(str).str.lower().isin(["inconclusive", "inadequate"])).sum()
            invalid_coords = df_internal[result_col].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"]).sum()

            cca_row = [
                formatted_date,   # Date
                description_val,  # Description
                read_type,        # Read Type (blank/default)
                "CCA",           # Network/System Type
                endpoint_type,    # Endpoint Type
                collectors,       # Collectors
                total_services,   # Total Services
                invalid_coords,              # Invalid Coordinates
                adequate_coverage,              # Adequate Coverage
                total_inconclusive,              # Inconclusive Coverage
                area_total,              # Area Total
                area_coverage,              # Area Coverage
                coverage_case     # Coverage Analysis Case
            ]
            tree.insert("", tk.END, values=cca_row)
        elif "Result" in df_internal.columns:
            adequate_coverage = (
                (df_internal["Result"].astype(str).str.lower() == "firstnet") |
                (df_internal["Result"].astype(str).str.lower() == "verizon") |
                (df_internal["Result"].astype(str).str.lower() == "t-mobile")|
                (df_internal["Result"].astype(str).str.lower() == "r900_cellular")
            ).sum()
            total_inconclusive = (df_internal["Result"].astype(str).str.lower() == "inconclusive").sum()
            cca_row = [
                formatted_date,   # Date
                description_val,  # Description
                read_type,        # Read Type (blank/default)
                "CCA",           # Network/System Type
                endpoint_type,    # Endpoint Type
                collectors,       # Collectors
                total_services,   # Total Services
                invalid_coords,              # Invalid Coordinates
                adequate_coverage,              # Adequate Coverage
                total_inconclusive,              # Inconclusive Coverage
                area_total,              # Area Total
                area_coverage,              # Area Coverage
                coverage_case     # Coverage Analysis Case
            ]
            tree.insert("", tk.END, values=cca_row)
        # else: skip row if neither column exists

    def senet_processing(self):
        # check each internal file for a column labeled "Predicted Coverage". If found, process for Senet.
        for internal_path in self.internal_files:
            try:
                df_internal = pd.read_csv(internal_path)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read internal CSV: {e}")
                return
            if "Predicted Coverage" not in df_internal.columns:
                messagebox.showerror("Error", f"'Predicted Coverage' column not found in {os.path.basename(internal_path)}. Cannot process for Senet.")
                return
            # If column 'Predicted Coverage' has values, "DR1" and "Yes" then the read type is Daily. Otherwise if there is no "Yes" then readtype is Hourly.
            read_type = "Hourly"
            if df_internal["Predicted Coverage"].astype(str).str.upper().str.contains("YES").any():
                read_type = "Daily" 

        return

    def on_submit(self):
        # Error check: Ensure at least one internal file is loaded
        if not self.internal_files:
            messagebox.showerror("Error", "No internal files loaded. Please add at least one internal CSV file before submitting.")
            return
        # Check if Senet button is selected if not check if Service area file was loaded. If not, show error.
        if self.senet_var.get() == 0 and self.service_area_df is None:
            messagebox.showerror("Error", "Please load a Service Area Excel file before submitting for Senet.")
            return
        # Remove any previous Accept/Back buttons from the logo frame (if it exists)
        if hasattr(self, 'logo_frame') and self.logo_frame.winfo_exists():
            for widget in self.logo_frame.winfo_children():
                if getattr(widget, '_di_action_button', False):
                    widget.destroy()
        # Remove the original logo frame before adding a new one
        if hasattr(self, 'logo_frame') and self.logo_frame.winfo_exists():
            self.logo_frame.destroy()
        # Recreate the logo frame on the right
        self.logo_frame = tk.Frame(self.main_area, bg="#d3d3d3", width=200)
        self.logo_frame.pack(side=tk.RIGHT, fill=tk.Y)
        self.logo_frame.pack_propagate(False)
        logo_path = os.path.join(os.path.dirname(__file__), "NeptuneLogo.png")
        try:
            if not hasattr(self, "_logo_img"):
                self._logo_img = tk.PhotoImage(file=logo_path)
            logo_label = tk.Label(self.logo_frame, image=self._logo_img, bg="#d3d3d3")
            logo_label.place(relx=0.5, rely=0.5, anchor="center")
            self._logo_label = logo_label
        except Exception as e:
            pass

        if self.senet_var.get() == 1:
            self.senet_processing()
        
        
        # Add Accept and Back buttons to the right side (logo_frame) after logo is created
        def on_accept():
            # Export the Data Import table to CSV in Z:/Propagations/DataImport or fallback to working directory
            di_dir = r"Z:\\Propagations\\DataImport"
            # Find the Data Import table's Treeview widget
            di_tree = None
            if hasattr(self, '_di_table_frame'):
                for widget in self._di_table_frame.winfo_children():
                    for subwidget in widget.winfo_children():
                        if isinstance(subwidget, ttk.Treeview):
                            di_tree = subwidget
                            break
                    if di_tree:
                        break
            if di_tree is None:
                messagebox.showerror("Error", "Could not find Data Import table to export.")
                return
            # Get data from the Data Import table
            columns = di_tree.cget("columns")
            data = [columns]
            for item in di_tree.get_children():
                row = di_tree.item(item, "values")
                data.append(row)
            # Build output filename
            internal_path = self.internal_files[0] if self.internal_files else "internal.csv"
            filename = os.path.splitext(os.path.basename(internal_path))[0]
            # Extract everything before and including the first YYMMDD date
            match = re.search(r'(.*?\d{6})', filename)
            if match:
                base = match.group(1)
            else:
                base = filename
            out_name = f"{base} Data Import.csv"
            out_path = os.path.join(di_dir, out_name)
            try:
                os.makedirs(di_dir, exist_ok=True)
                with open(out_path, "w", newline='', encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerows(data)
                messagebox.showinfo("Accepted", f"Data Import Data exported to {out_path}")
            except Exception:
                # Fallback to working directory
                out_path = os.path.join(os.getcwd(), out_name)
                try:
                    with open(out_path, "w", newline='', encoding="utf-8") as f:
                        writer = csv.writer(f)
                        writer.writerows(data)
                    messagebox.showinfo("Accepted", f"Data Import Data exported to {out_path}")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to export Data Import Data: {e}")

        def on_back():
            # Hide Data Import table and show main table again
            if hasattr(self, '_di_table_frame') and self._di_table_frame.winfo_exists():
                self._di_table_frame.pack_forget()
            if hasattr(self, 'table_frame') and self.table_frame.winfo_exists():
                parent = self.table_frame.master
                parent.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            # Remove Accept/Back buttons frame
            if hasattr(self, '_di_btns_frame') and self._di_btns_frame.winfo_exists():
                self._di_btns_frame.destroy()
            # Extra: Remove any button frames or widgets in logo_frame that are not the logo label
            if hasattr(self, 'logo_frame') and self.logo_frame.winfo_exists():
                for widget in self.logo_frame.winfo_children():
                    # Only destroy if it's a Frame or Button and not the logo label
                    if widget is not getattr(self, '_logo_label', None):
                        widget.destroy()

        # Create a frame to hold the buttons, placed below the logo and centered
        btns_frame = tk.Frame(self.logo_frame, bg="#d3d3d3")
        btns_frame.place(relx=0.5, rely=0.65, anchor="n")  # 0.65 = below center
        self._di_btns_frame = btns_frame  # Save reference for removal

        accept_btn = tk.Button(btns_frame, text="Accept", command=on_accept, font=("Segoe UI", 10, "bold"), bg="#2e7da8", fg="white", activebackground="#246688", activeforeground="white", bd=0, relief=tk.FLAT, padx=16, pady=8)
        accept_btn._di_action_button = True
        accept_btn.pack(side=tk.TOP, pady=(0, 12), fill=tk.X)
        back_btn = tk.Button(btns_frame, text="Back", command=on_back, font=("Segoe UI", 10, "bold"), bg="#a8a8a8", fg="white", activebackground="#888888", activeforeground="white", bd=0, relief=tk.FLAT, padx=16, pady=8)
        back_btn._di_action_button = True
        back_btn.pack(side=tk.TOP, fill=tk.X)

        # Take the internal files.
        internal_path = self.internal_files[0]
        try:
            df_internal = pd.read_csv(internal_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read internal CSV: {e}")
            return

        # Hide the main table frame (parent of self.table_frame) but keep the logo frame visible on the right
        if hasattr(self, 'table_frame') and self.table_frame.winfo_ismapped():
            parent = self.table_frame.master
            parent.pack_forget()

        # Remove the original logo frame before adding a new one
        if hasattr(self, 'logo_frame') and self.logo_frame.winfo_exists():
            self.logo_frame.destroy()
        # Recreate the logo frame on the right
        self.logo_frame = tk.Frame(self.main_area, bg="#d3d3d3", width=200)
        self.logo_frame.pack(side=tk.RIGHT, fill=tk.Y)
        self.logo_frame.pack_propagate(False)
        logo_path = os.path.join(os.path.dirname(__file__), "NeptuneLogo.png")
        try:
            if not hasattr(self, "_logo_img"):
                self._logo_img = tk.PhotoImage(file=logo_path)
            logo_label = tk.Label(self.logo_frame, image=self._logo_img, bg="#d3d3d3")
            logo_label.place(relx=0.5, rely=0.5, anchor="center")
            self._logo_label = logo_label
        except Exception as e:
            pass

        # Create a frame to hold the buttons, placed below the logo and centered
        btns_frame = tk.Frame(self.logo_frame, bg="#d3d3d3")
        btns_frame.place(relx=0.5, rely=0.65, anchor="n")  # 0.65 = below center

        accept_btn = tk.Button(btns_frame, text="Accept", command=on_accept, font=("Segoe UI", 10, "bold"), bg="#2e7da8", fg="white", activebackground="#246688", activeforeground="white", bd=0, relief=tk.FLAT, padx=16, pady=8)
        accept_btn._di_action_button = True
        accept_btn.pack(side=tk.TOP, pady=(0, 12), fill=tk.X)
        back_btn = tk.Button(btns_frame, text="Back", command=on_back, font=("Segoe UI", 10, "bold"), bg="#a8a8a8", fg="white", activebackground="#888888", activeforeground="white", bd=0, relief=tk.FLAT, padx=16, pady=8)
        back_btn._di_action_button = True
        back_btn.pack(side=tk.TOP, fill=tk.X)

        # If a Data import table already exists, clear it; otherwise, create it
        if not hasattr(self, '_di_table_frame') or not self._di_table_frame.winfo_exists():
            self._di_table_frame = tk.Frame(self.main_area, bg="#d3d3d3")
            self._di_table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=(10, 30))
        else:
            for widget in self._di_table_frame.winfo_children():
                widget.destroy()
            self._di_table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=(10, 30))
        # New title for Data Import table
        title = tk.Label(self._di_table_frame, text="DATA Import", bg="#d3d3d3", fg="#222", font=("Segoe UI", 14, "bold"))
        title.pack(pady=(16, 4))

        # --- Senet logic: If 'Predicted Coverage' column exists, process as Senet ---
        # Check all internal files for Senet and create a row for each
        senet_files = []
        for internal_path in self.internal_files:
            try:
                df_internal = pd.read_csv(internal_path)
                if "Result" in df_internal.columns:
                    df_internal["Result"] = df_internal["Result"].replace("R900_ Wall External LoRaWAN", "R900_Wall External LoRaWAN")
                if "Result" in df_internal.columns:
                    df_internal["Result"] = df_internal["Result"].replace("R900_ Wall External LoRaWAN", "R900_Wall External LoRaWAN")
            except Exception:
                continue
            if "Predicted Coverage" in df_internal.columns:
                senet_files.append((internal_path, df_internal))
        if senet_files:
            columns = ["Date", "Description", "Read Type", "Network/System Type", "Endpoint Type", "Collectors", "Total Services", "Invalid Coordinates", "Adequate Coverage", "Inconclusive Coverage", "Area Total", "Area Coverage", "Coverage Analysis Case"]
            style = ttk.Style()
            style.configure("DI.Treeview", borderwidth=1, relief="solid", rowheight=24)
            style.configure("DI.Treeview.Heading", borderwidth=1, relief="solid")
            style.map("DI.Treeview", background=[('selected', '#cce6ff')])
            style.layout("DI.Treeview", [
                ('Treeview.field', {'sticky': 'nswe', 'children': [
                    ('Treeview.padding', {'sticky': 'nswe', 'children': [
                        ('Treeview.treearea', {'sticky': 'nswe'})
                    ]})
                ]})
            ])
            di_table_inner = tk.Frame(self._di_table_frame, bg="#d3d3d3")
            di_table_inner.pack(fill=tk.BOTH, expand=True)
            tree = ttk.Treeview(di_table_inner, columns=columns, show="headings", style="DI.Treeview")
            tree.tag_configure('oddrow', background='#f2f2f2')
            tree.tag_configure('evenrow', background='#ffffff')
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=120, anchor=tk.CENTER)
            vsb = ttk.Scrollbar(di_table_inner, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(di_table_inner, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            di_table_inner.grid_columnconfigure(0, weight=1)
            di_table_inner.grid_rowconfigure(0, weight=1)

            def insert_striped_row(idx, vals):
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                tree.insert("", tk.END, values=vals, tags=(tag,))

            for file_idx, (internal_path, df_internal) in enumerate(senet_files):
                filename = os.path.basename(internal_path)
                date_match = re.search(r'(\d{6})', filename)
                if date_match:
                    date_str = date_match.group(1)
                    month = date_str[2:4]
                    day = date_str[4:6]
                    year = "20" + date_str[0:2]
                    formatted_date = f"{month}/{day}/{year}"
                else:
                    formatted_date = "01/01/2000"  # default date if not found

                senet_count = 0
                total_services = len(df_internal)
                invalid_coords = 0
                adequate_coverage = 0
                inconcl_cov = 0
                has_daily = False
                for _, row in df_internal.iterrows():
                    result_val = str(row.get("Result", ""))
                    pred_cov = str(row.get("Predicted Coverage", "")).strip().lower()
                    # Set endpoint_type based on Result value
                    if "Pit LoRaWAN" in result_val:
                        endpoint_type = "Pit"
                    elif "Wall External LoRaWAN" in result_val:
                        endpoint_type = "Wall External"
                    elif "Wall Inside" in result_val:
                        endpoint_type = "Wall Inside"
                    else:
                        endpoint_type = ""
                    if "lorawan" in result_val.lower():
                        senet_count += 1
                    if pred_cov == "yes" and "lorawan" in result_val.lower():
                        has_daily = True
                    result_upper = result_val.strip().upper()
                    if result_upper in ("INVALID LOCATION", "INVALID_LOCATION"):
                        invalid_coords += 1
                    elif result_upper == "INCONCLUSIVE":
                        inconcl_cov += 1
                    else:
                        adequate_coverage += 1
                read_type = "Daily" if has_daily else "Hourly"
                row_vals = [
                    formatted_date,
                    "Senet NaaS",
                    read_type,
                    "R900_LoRaNaaS",
                    endpoint_type,
                    "-",  # Collectors not relevant for Senet
                    total_services,
                    invalid_coords,
                    adequate_coverage,
                    inconcl_cov,
                    0 if "Area Total" not in df_internal.columns else df_internal["Area Total"].iloc[0] if not df_internal.empty else "",
                    0 if "Area Coverage" not in df_internal.columns else df_internal["Area Coverage"].iloc[0] if not df_internal.empty else "",
                    self.coverage_entry.get().strip()
                ]
                insert_striped_row(file_idx, row_vals)

                # --- Add cellular provider rows and CCA row if cellular columns exist ---
                cellular_cols = ["LTE-M_-100", "CAT-M", "L700-CATM", "L1900-CATM"]
                has_cellular_data = any(
                    col in df_internal.columns and df_internal[col].notna().any() and (df_internal[col] != '').any()
                    for col in cellular_cols
                )
                if has_cellular_data:
                    providers = []
                    if "LTE-M_-100" in df_internal.columns:
                        providers.append("FirstNet")
                    if "CAT-M" in df_internal.columns:
                        providers.append("Verizon")
                    if "L700-CATM" in df_internal.columns or "L1900-CATM" in df_internal.columns:
                        providers.append("T-Mobile")
                    for provider in providers:
                        if provider == "FirstNet":
                            adeq_cov = df_internal[(df_internal["LTE-M_-100"].astype(str).str.upper() == "Y") | (df_internal["LTE-M_-100"].astype(str).str.upper() == "YES")].shape[0]
                        elif provider == "Verizon":
                            adeq_cov = df_internal[(pd.to_numeric(df_internal["CAT-M"], errors='coerce') >= -100) & (pd.to_numeric(df_internal["CAT-M"], errors='coerce') != 0)].shape[0]
                        elif provider == "T-Mobile":
                            adeq_cov = df_internal[(df_internal["L700-CATM"].astype(str).str.upper().isin(["IBC", "IBR"])) | (df_internal["L1900-CATM"].astype(str).str.upper().isin(["IBC", "IBR"]))].shape[0]
                        else:
                            adeq_cov = 0
                        # invalid_coords_cell = df_internal[df_internal["Result"].astype(str).str.upper() == "INVALID LOCATION"].shape[0]
                        invalid_coords_cell = df_internal[df_internal["Result"].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"])].shape[0]
                        inconcl_cov_cell = total_services - adeq_cov - invalid_coords_cell
                        row_vals_cell = [
                            formatted_date,
                            provider,
                            "",
                            provider,
                            "R900_Cellular",
                            0,
                            total_services,
                            invalid_coords_cell,
                            adeq_cov,
                            inconcl_cov_cell,
                            0 if "Area Total" not in df_internal.columns else df_internal["Area Total"].iloc[0] if not df_internal.empty else "",
                            0 if "Area Coverage" not in df_internal.columns else df_internal["Area Coverage"].iloc[0] if not df_internal.empty else "",
                            self.coverage_entry.get().strip()
                        ]
                        insert_striped_row(file_idx + 1, row_vals_cell)
                    # Add CCA summary row
                    adequate_mask = pd.Series([False] * len(df_internal))
                    if "LTE-M_-100" in df_internal.columns:
                        adequate_mask = adequate_mask | (df_internal["LTE-M_-100"].astype(str).str.upper().isin(["Y", "YES"]))
                    if "CAT-M" in df_internal.columns:
                        adequate_mask = adequate_mask | ((pd.to_numeric(df_internal["CAT-M"], errors='coerce') >= -100) & (pd.to_numeric(df_internal["CAT-M"], errors='coerce') != 0))
                    if "L700-CATM" in df_internal.columns:
                        adequate_mask = adequate_mask | (df_internal["L700-CATM"].astype(str).str.upper().isin(["IBC", "IBR"]))
                    if "L1900-CATM" in df_internal.columns:
                        adequate_mask = adequate_mask | (df_internal["L1900-CATM"].astype(str).str.upper().isin(["IBC", "IBR"]))
                    adequate_coverage_cca = adequate_mask.sum()
                    invalid_mask = ("Result" in df_internal.columns) and (df_internal["Result"].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"]))
                    if isinstance(invalid_mask, pd.Series):
                        total_inconclusive = (~adequate_mask & ~invalid_mask).sum()
                        invalid_coords_cca = invalid_mask.sum()
                    else:
                        total_inconclusive = (~adequate_mask).sum()
                        invalid_coords_cca = 0
                    cca_row = [
                        formatted_date,
                        "CCA",
                        "",
                        "CCA",
                        "R900_Cellular",
                        0,
                        total_services,
                        invalid_coords_cca,
                        adequate_coverage_cca,
                        total_inconclusive,
                        0 if "Area Total" not in df_internal.columns else df_internal["Area Total"].iloc[0] if not df_internal.empty else "",
                        0 if "Area Coverage" not in df_internal.columns else df_internal["Area Coverage"].iloc[0] if not df_internal.empty else "",
                        self.coverage_entry.get().strip()
                    ]
                    insert_striped_row(file_idx + 2, cca_row)
            return  # Done with Senet, skip rest of logic

        columns = ["Date", "Description", "Read Type", "Network/System Type", "Endpoint Type", "Collectors", "Total Services", "Invalid Coordinates", "Adequate Coverage", "Inconclusive Coverage", "Area Total", "Area Coverage", "Coverage Analysis Case"]
        style = ttk.Style()
        style.configure("DI.Treeview", borderwidth=1, relief="solid", rowheight=24)
        style.configure("DI.Treeview.Heading", borderwidth=1, relief="solid")
        style.map("DI.Treeview", background=[('selected', '#cce6ff')])
        style.layout("DI.Treeview", [
            ('Treeview.field', {'sticky': 'nswe', 'children': [
                ('Treeview.padding', {'sticky': 'nswe', 'children': [
                    ('Treeview.treearea', {'sticky': 'nswe'})
                ]})
            ]})
        ])
        # Use a frame for the table and scrollbars
        di_table_inner = tk.Frame(self._di_table_frame, bg="#d3d3d3")
        di_table_inner.pack(fill=tk.BOTH, expand=True)
        tree = ttk.Treeview(di_table_inner, columns=columns, show="headings", style="DI.Treeview")
        # Add striped row tags for alternating colors (after tree is created)
        tree.tag_configure('oddrow', background='#f2f2f2')
        tree.tag_configure('evenrow', background='#ffffff')
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor=tk.CENTER)
        vsb = ttk.Scrollbar(di_table_inner, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(di_table_inner, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        di_table_inner.grid_columnconfigure(0, weight=1)
        di_table_inner.grid_rowconfigure(0, weight=1)

        # --- Add Delete Row button below the table ---
        def delete_selected_rows():
            selected_items = tree.selection()
            if not selected_items:
                messagebox.showinfo("Delete Row", "No row selected. Please select a row to delete.")
                return
            for item in selected_items:
                tree.delete(item)

        # --- Add right-click context menu for deleting rows ---
        menu = tk.Menu(tree, tearoff=0)
        menu.add_command(label="Delete Row", command=delete_selected_rows)

        def on_right_click(event):
            # Select row under mouse if not already selected
            row_id = tree.identify_row(event.y)
            if row_id:
                if row_id not in tree.selection():
                    tree.selection_set(row_id)
                menu.tk_popup(event.x_root, event.y_root)
            else:
                # If not clicking on a row, clear selection
                tree.selection_remove(tree.selection())

        tree.bind("<Button-3>", on_right_click)

        btn_frame = tk.Frame(self._di_table_frame, bg="#d3d3d3")
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10), anchor="w")
        del_btn = tk.Button(btn_frame, text="Delete Row", command=delete_selected_rows, font=("Segoe UI", 10, "bold"), bg="#d9534f", fg="white", activebackground="#b52a1d", activeforeground="white", bd=0, relief=tk.FLAT, padx=12, pady=6)
        del_btn.pack(side=tk.LEFT, pady=2)

        # --- Make Data Import table cells editable on click ---
        def on_di_cell_click(event):
            region = tree.identify("region", event.x, event.y)
            if region != "cell":
                return
            row_id = tree.identify_row(event.y)
            col_id = tree.identify_column(event.x)  # '#1'..'#N'
            col_index = int(col_id.replace('#', '')) - 1
            bbox = tree.bbox(row_id, col_id)
            if not bbox:
                return
            x, y, w, h = bbox
            current = tree.set(row_id, columns[col_index])
            # Destroy any previous Entry widget
            if hasattr(tree, '_active_entry') and tree._active_entry is not None:
                try:
                    tree._active_entry.destroy()
                except Exception:
                    pass
                tree._active_entry = None
            entry = tk.Entry(tree)
            tree._active_entry = entry
            entry.place(x=x, y=y, width=w, height=h)
            entry.insert(0, current)
            entry.focus_set()

            def finish(_evt=None):
                tree.set(row_id, columns[col_index], entry.get())
                entry.destroy()
                tree._active_entry = None

            entry.bind("<Return>", finish)
            entry.bind("<FocusOut>", lambda e: finish())

        tree.bind("<Button-1>", on_di_cell_click)

        filename = os.path.basename(internal_path)
        date_match = re.search(r'(\d{6})', filename)
        if date_match:
            date_str = date_match.group(1)
            month = date_str[2:4]
            day = date_str[4:6]
            year = "20" + date_str[0:2]
            formatted_date = f"{month}/{day}/{year}"
        else:
            formatted_date = "01/01/2000"  # default date if not found

       # For each row in the service area dataframe, add a row to the Data Import table. we will look at the MIU Type column and match it to the value in the result column of the mapped internal file. 
        # When inserting rows, alternate tags for striped effect
        def insert_striped_row(idx, vals):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            tree.insert("", tk.END, values=vals, tags=(tag,))

        striped_row_idx = 0
        
        miu_to_result = {
            "R900iv3 Basement": "R900_v3 Basement",
            "R900v3 Inside Wall": "R900_v3 Wall Inside",
            "R900v3 Wall Ext": "R900_v3 Wall External",
            "R900v3 Pit": "R900_v3 Pit",
            "R900iv4 Basement": "R900_v4 Basement",
            "R900v4 Inside Wall": "R900_v4 Wall Inside",
            "R900v4 Inside": "R900_v4 Wall Inside",
            "R900v4 Wall Ext": "R900_v4 Wall External",
            "R900v4 Wall": "R900_v4 Wall External",
            "R900v4 Pit": "R900_v4 Pit",
            "R900v5 Inside Wall": [
                "R900_v5 Wall Inside LoRaWAN",
                "R900_Wall Inside LoRaWAN",
                "R900_ Wall Inside LoRaWAN"
            ],
            "R900v5 Wall Ext": [
                "R900_v5 Wall External LoRaWAN",
                "R900_Wall External LoRaWAN",
                "R900_ Wall External LoRaWAN"
            ],
            "R900v5 Pit": [
                "R900_v5 Pit LoRaWAN",
                "R900_Pit LoRaWAN",
                "R900_ Pit LoRaWAN"]
        }

        # Get area_total from cell I1 and build a map->area_coverage lookup from the service area Excel file
        area_total = ""
        area_coverage_lookup = {}
        try:
            if self.service_area_path:
                import openpyxl
                wb = openpyxl.load_workbook(self.service_area_path, data_only=True)
                ws = wb.active
                area_total = ws["I1"].value if ws["I1"].value is not None else ""
                # Build lookup: map value (col A) -> area_coverage (col H)
                for row in ws.iter_rows(min_row=3):  # skip header rows (1,2)
                    map_cell = row[0].value
                    area_cov_cell = row[7].value if len(row) > 7 else None
                    if map_cell is not None:
                        area_coverage_lookup[str(map_cell).strip()] = area_cov_cell if area_cov_cell is not None else ""
        except Exception:
            area_total = ""
            area_coverage_lookup = {}

        mapped_internal_indices = []
        used_internal_files = set()
        # --- Collect all unique collector IDs across all mapped internal files ---
        global_unique_collectors = set()
        collectors_by_internal = {}
        suffixes = ("_B", "_W", "_P", "_I")
        def strip_collector_suffix(collector_id):
            for suf in suffixes:
                if collector_id.endswith(suf):
                    return collector_id[:-len(suf)]
            return collector_id
        
        already_processed = set()
        # First, collect all unique collector IDs from all mapped internal files
        for item in self.tree.get_children():
            vals = list(self.tree.item(item, "values"))
            if len(vals) < 5 or not vals[4]:
                continue
            internal_idx = ord(vals[4]) - ord('A') if vals[4].isalpha() else None
            if internal_idx is None or internal_idx < 0 or internal_idx >= len(self.internal_files):
                continue
            mapped_internal_indices.append(internal_idx)
            used_internal_files.add(internal_idx)

            if internal_idx in already_processed:
                continue
            already_processed.add(internal_idx)

            internal_path = self.internal_files[internal_idx]
            try:
                df_internal_row = pd.read_csv(internal_path)
                if "Result" in df_internal_row.columns:
                    df_internal_row["Result"] = df_internal_row["Result"].replace("R900_ Wall External LoRaWAN", "R900_Wall External LoRaWAN")
                # Add all unique collector IDs from this file, stripping suffixes
                
                local_collectors = set()

                for col in df_internal_row.columns:
                    col_str = str(col)
                    if col_str.startswith(('3', '4', '5')):
                        base_id = col_str[1:]
                        base_id = strip_collector_suffix(base_id)
                        local_collectors.add(base_id)

                collectors_by_internal[internal_idx] = local_collectors
                global_unique_collectors |= local_collectors

            except Exception:
                continue

        # Now, for each row, use the global unique collector count
        for item in self.tree.get_children():
            vals = list(self.tree.item(item, "values"))
            if len(vals) < 5:
                continue
            if not vals[4]:
                continue
            internal_idx = ord(vals[4]) - ord('A') if vals[4].isalpha() else None
            if internal_idx is None or internal_idx < 0 or internal_idx >= len(self.internal_files):
                continue
            internal_path = self.internal_files[internal_idx]
            try:
                df_internal_row = pd.read_csv(internal_path)
                if "Result" in df_internal_row.columns:
                    df_internal_row["Result"] = df_internal_row["Result"].replace("R900_ Wall External LoRaWAN", "R900_Wall External LoRaWAN")
            except Exception:
                continue
            desc_val = vals[1]
            read_type = vals[3]
            # Set endpoint_type based on MIU Type value
            if "R900v5" in vals[2]:
                miu_type = "R900_LoRaWAN"
            elif "R900v4" in vals[2]:
                miu_type = "R900_v4"
            elif "R900v3" in vals[2]:
                miu_type = "R900_v3"
            else:
                miu_type = vals[2].split()[0]
            if "LoRaWAN" in vals[2]:
                if "Basement" in vals[2]:
                    endpoint_type = "Basement LoRaWAN"
                elif "Inside Wall" in vals[2]:
                    endpoint_type = "Wall Inside LoRaWAN"
                elif "Wall Ext" in vals[2]:
                    endpoint_type = "Wall External LoRaWAN"
                elif "Pit" in vals[2]:
                    endpoint_type = "Pit LoRaWAN"
                else:
                    endpoint_type = "LoRaWAN"
            else:
                if "Basement" in vals[2]:
                    endpoint_type = "Basement"
                elif "Inside Wall" in vals[2] or "Inside" in vals[2]:
                    endpoint_type = "Wall Inside"
                elif "Wall Ext" in vals[2] or "Wall" in vals[2]:
                    endpoint_type = "Wall External"
                elif "Pit" in vals[2]:
                    endpoint_type = "Pit"
                else:
                    endpoint_type = ""
            result_val = miu_to_result.get(vals[2], "")
            if isinstance(result_val, list):
                matched_services = df_internal_row[df_internal_row["Result"].astype(str).str.upper().isin([v.upper() for v in result_val])]
            elif result_val:
                matched_services = df_internal_row[df_internal_row["Result"].astype(str).str.upper() == result_val.upper()]
            else:
                matched_services = pd.DataFrame()
            total_services = len(df_internal_row)
            invalid_coords = df_internal_row[df_internal_row["Result"].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"])].shape[0]
            miu_hierarchy = [
                "R900v5 Wall Ext", "R900v5 Pit", "R900v5 Inside Wall",
                "R900v4 Wall", "R900v4 Wall Ext", "R900v4 Pit", "R900v4 Inside Wall", "R900iv4 Basement",
                "R900v3 Wall Ext", "R900v3 Pit", "R900v3 Inside Wall", "R900iv3 Basement"
            ]
            miu_to_result = {
                "R900iv3 Basement": "R900_v3 Basement",
                "R900v3 Inside Wall": "R900_v3 Wall Inside",
                "R900v3 Wall Ext": "R900_v3 Wall External",
                "R900v3 Pit": "R900_v3 Pit",
                "R900iv4 Basement": "R900_v4 Basement",
                "R900v4 Inside": "R900_v4 Wall Inside",
                "R900v4 Inside Wall": "R900_v4 Wall Inside",
                "R900v4 Wall": "R900_v4 Wall External",
                "R900v4 Wall Ext": "R900_v4 Wall External",
                "R900v4 Pit": "R900_v4 Pit",
                "R900v5 Inside Wall": [
                    "R900_v5 Wall Inside LoRaWAN",
                    "R900_Wall Inside LoRaWAN",
                    "R900_ Wall Inside LoRaWAN"
                    ],
                "R900v5 Wall Ext": [
                    "R900_v5 Wall External LoRaWAN",
                    "R900_Wall External LoRaWAN",
                    "R900_ Wall External LoRaWAN"
                    ],
                "R900v5 Pit": [
                    "R900_v5 Pit LoRaWAN",
                    "R900_Pit LoRaWAN",
                    "R900_ Pit LoRaWAN"
                    ]
            }
            miu_type_full = vals[2]
            if miu_type_full in miu_hierarchy:
                idx = miu_hierarchy.index(miu_type_full)
                mius_to_sum = miu_hierarchy[idx:]
            else:
                mius_to_sum = [miu_type_full]
            result_values = [miu_to_result.get(miu, miu) for miu in mius_to_sum]
            adequate_coverage = 0
            for result_val in result_values:
                if isinstance(result_val, list):
                    ms = df_internal_row[df_internal_row["Result"].astype(str).str.upper().isin([v.upper() for v in result_val])]
                else:
                    ms = df_internal_row[df_internal_row["Result"].astype(str).str.upper() == result_val.upper()]
                adequate_coverage += ms[(~ms["Result"].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"])) & (ms["Result"].astype(str).str.upper() != "INCONCLUSIVE")].shape[0]
            inconcl_cov = total_services - adequate_coverage - invalid_coords
            # Use the global unique collector count for all rows
            collectors = len(collectors_by_internal.get(internal_idx, set()))  # Optionally, use per-file collector count instead of global
            
            # collectors = len(global_unique_collectors)
            map_val = vals[0]
            area_coverage = area_coverage_lookup.get(str(map_val).strip(), "")
            coverage_case = self.coverage_entry.get().strip()
            row_vals = [
                formatted_date,
                desc_val,
                read_type,
                miu_type,
                endpoint_type,
                collectors,
                total_services,
                invalid_coords,
                adequate_coverage,
                inconcl_cov,
                area_total,
                area_coverage,
                coverage_case
            ]
            insert_striped_row(striped_row_idx, row_vals)
            striped_row_idx += 1

        # --- Append Tailored row for each internal file used, if applicable ---
        for idx in sorted(used_internal_files):
            internal_path = self.internal_files[idx]
            try:
                df_internal = pd.read_csv(internal_path)
            except Exception:
                continue
            exclude_results = {"FIRSTNET", "VERIZON", "T-MOBILE", "INCONCLUSIVE", "INVALID LOCATION", "INVALID_LOCATION"}
            miu_types = set(
                str(x).strip() for x in df_internal["Result"].dropna().unique()
                if str(x).strip().upper() not in exclude_results
            ) if "Result" in df_internal.columns else set()
            is_cellular = any(col in df_internal.columns for col in ["LTE-M_-100", "CAT-M", "L700-CATM", "L1900-CATM"])
            # Only add tailored row if NOT (only 1 miu and not cellular)
            if len(miu_types) == 1 and not is_cellular:
                continue  # skip tailored row if only 1 miu and not cellular
            # Otherwise, add tailored row if is_cellular or more than 1 miu
            if not (is_cellular or len(miu_types) > 1):
                continue
            endpoint_type = ""
            # Use the same global_unique_collectors logic for tailored row
            collectors = len(global_unique_collectors)
            total_services = len(df_internal)
            area_total_val = area_total
            area_coverage_val = ""
            coverage_case_val = self.coverage_entry.get().strip()
            read_type_val = ""
            for item in self.tree.get_children():
                vals = list(self.tree.item(item, "values"))
                if len(vals) >= 5:
                    internal_idx = ord(vals[4]) - ord('A') if vals[4] and vals[4].isalpha() else None
                    if internal_idx == idx:
                        read_type_val = vals[3]
                        break
            tailored_services = df_internal[~df_internal["MIU Type"].astype(str).str.contains("LTE|CAT|L700|L1900", case=False, na=False)] if "MIU Type" in df_internal.columns else df_internal
            if "Result" in tailored_services.columns:
                invalid_coords = tailored_services["Result"].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"]).sum()
                adequate_coverage = (~tailored_services["Result"].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"]) & (tailored_services["Result"].astype(str).str.upper() != "INCONCLUSIVE")).sum()
            else:
                invalid_coords = 0
                adequate_coverage = 0
            inconcl_cov = total_services - adequate_coverage - invalid_coords
            row_vals = [
                formatted_date,
                "Tailored",
                read_type_val,
                "Tailored",
                endpoint_type,
                collectors,
                total_services,
                invalid_coords,
                adequate_coverage,
                inconcl_cov,
                area_total_val,
                area_coverage_val,
                coverage_case_val
            ]
            insert_striped_row(striped_row_idx, row_vals)
            striped_row_idx += 1

        # --- Append CCA summary rows as if CCA button was pressed ---
        # Use the first mapped internal file for CCA summary
        cca_idx = mapped_internal_indices[0] if mapped_internal_indices else None
        if cca_idx is not None and cca_idx < len(self.internal_files):
            cca_internal_path = self.internal_files[cca_idx]
            try:
                df_internal = pd.read_csv(cca_internal_path)
                if "Result" in df_internal.columns:
                    df_internal["Result"] = df_internal["Result"].replace("R900_ Wall External LoRaWAN", "R900_Wall External LoRaWAN")
            except Exception:
                df_internal = None
            if df_internal is not None:
                # Only add CCA row if there are cellular outputs (at least one of the columns exists and has non-empty data)
                cellular_cols = ["LTE-M_-100", "CAT-M", "L700-CATM", "L1900-CATM"]
                has_cellular_data = any(
                    col in df_internal.columns and df_internal[col].notna().any() and (df_internal[col] != '').any()
                    for col in cellular_cols
                )
                if not has_cellular_data:
                    pass  # do not add CCA row
                else:
                    endpoint_type = "R900_Cellular"
                    collectors = 0
                    total_services = len(df_internal)
                    area_total_val = area_total
                    area_coverage_val = ""
                    coverage_case_val = self.coverage_entry.get().strip()
                    read_type_val = ""
                    providers = []
                    if "LTE-M_-100" in df_internal.columns:
                        providers.append("FirstNet")
                    if "CAT-M" in df_internal.columns:
                        providers.append("Verizon")
                    if "L700-CATM" in df_internal.columns or "L1900-CATM" in df_internal.columns:
                        providers.append("T-Mobile")
                    for provider in providers:
                        if provider == "FirstNet":
                            adeq_cov = df_internal[(df_internal["LTE-M_-100"].astype(str).str.upper() == "Y") | (df_internal["LTE-M_-100"].astype(str).str.upper() == "YES")].shape[0]
                        elif provider == "Verizon":
                            adeq_cov = df_internal[(pd.to_numeric(df_internal["CAT-M"], errors='coerce') >= -100) & (pd.to_numeric(df_internal["CAT-M"], errors='coerce') != 0)].shape[0]
                        elif provider == "T-Mobile":
                            adeq_cov = df_internal[(df_internal["L700-CATM"].astype(str).str.upper().isin(["IBC", "IBR"])) | (df_internal["L1900-CATM"].astype(str).str.upper().isin(["IBC", "IBR"]))].shape[0]
                        else:
                            adeq_cov = 0
                        invalid_coords = df_internal[df_internal["Result"].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"])].shape[0]
                        inconcl_cov = total_services - adeq_cov - invalid_coords
                        # Find the first mapped service area row for this internal file to get the read type
                        read_type_val = ""
                        for item in self.tree.get_children():
                            vals = list(self.tree.item(item, "values"))
                            if len(vals) >= 5:
                                internal_idx = ord(vals[4]) - ord('A') if vals[4] and vals[4].isalpha() else None
                                if internal_idx == cca_idx:
                                    read_type_val = vals[3]
                                    break
                        row_vals = [
                            formatted_date,
                            provider,
                            "",
                            provider,
                            endpoint_type,
                            collectors,
                            total_services,
                            invalid_coords,
                            adeq_cov,
                            inconcl_cov,
                            area_total_val,
                            area_coverage_val,
                            coverage_case_val
                        ]
                        insert_striped_row(striped_row_idx, row_vals)
                        striped_row_idx += 1
                    # Add extra row with Description and Network/System Type as 'CCA'
                    adequate_mask = pd.Series([False] * len(df_internal))
                    if "LTE-M_-100" in df_internal.columns:
                        adequate_mask = adequate_mask | (df_internal["LTE-M_-100"].astype(str).str.upper().isin(["Y", "YES"]))
                    if "CAT-M" in df_internal.columns:
                        adequate_mask = adequate_mask | ((pd.to_numeric(df_internal["CAT-M"], errors='coerce') >= -100) & (pd.to_numeric(df_internal["CAT-M"], errors='coerce') != 0))
                    if "L700-CATM" in df_internal.columns:
                        adequate_mask = adequate_mask | (df_internal["L700-CATM"].astype(str).str.upper().isin(["IBC", "IBR"]))
                    if "L1900-CATM" in df_internal.columns:
                        adequate_mask = adequate_mask | (df_internal["L1900-CATM"].astype(str).str.upper().isin(["IBC", "IBR"]))
                    adequate_coverage = adequate_mask.sum()
                    invalid_mask = ("Result" in df_internal.columns) and (df_internal["Result"].astype(str).str.upper().isin(["INVALID LOCATION", "INVALID_LOCATION"]))
                    if isinstance(invalid_mask, pd.Series):
                        total_inconclusive = (~adequate_mask & ~invalid_mask).sum()
                        invalid_coords = invalid_mask.sum()
                    else:
                        total_inconclusive = (~adequate_mask).sum()
                        invalid_coords = 0
                    read_type_val = ""
                    for item in self.tree.get_children():
                        vals = list(self.tree.item(item, "values"))
                        if len(vals) >= 5:
                            internal_idx = ord(vals[4]) - ord('A') if vals[4] and vals[4].isalpha() else None
                            if internal_idx == cca_idx:
                                read_type_val = vals[3]
                                break
                    cca_row = [
                        formatted_date,   # Date
                        "CCA",           # Description
                        "",        # Read Type (from mapped service area row)
                        "CCA",           # Network/System Type
                        endpoint_type,    # Endpoint Type
                        0,       # Collectors
                        total_services,   # Total Services
                        invalid_coords,              # Invalid Coordinates
                        adequate_coverage,              # Adequate Coverage
                        total_inconclusive,              # Inconclusive Coverage
                        area_total_val,              # Area Total
                        area_coverage_val,              # Area Coverage
                        coverage_case_val     # Coverage Analysis Case
                    ]
                    tree.insert("", tk.END, values=cca_row)
       

    

if __name__ == "__main__":
    app = DataImportApp()
    app.mainloop()
